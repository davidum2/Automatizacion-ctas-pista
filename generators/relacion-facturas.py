import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def create_relacion_de_facturas_excel(data, output_dir, monto):
    """
    Crea un archivo Excel con la relación de facturas.
    
    Args:
        data (dict): Datos de la factura
        output_dir (str): Directorio donde se guardará el archivo
        monto (float): Monto de la factura
        
    Returns:
        str: Ruta al archivo Excel generado
    """
    try:
        # Crear un nuevo libro de trabajo
        wb = Workbook()
        ws = wb.active
        ws.title = "Relación de Facturas"
        
        # Estilos
        titulo_font = Font(name='Arial', size=14, bold=True)
        header_font = Font(name='Arial', size=12, bold=True)
        data_font = Font(name='Arial', size=11)
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Relleno para encabezados
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Título
        ws.merge_cells('A1:G1')
        cell = ws['A1']
        cell.value = "RELACIÓN DE FACTURAS"
        cell.font = titulo_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Información general
        ws['A3'] = "Partida:"
        ws['B3'] = data['No_partida']
        ws['A4'] = "Descripción:"
        ws['B4'] = data['Descripcion_partida']
        ws['A5'] = "Mes:"
        ws['B5'] = data['Mes']
        ws['A6'] = "Año:"
        ws['B6'] = datetime.now().year
        
        # Aplicar formato a la información general
        for cell in ws['A3:A6']:
            cell[0].font = header_font
            cell[0].alignment = Alignment(horizontal='right')
        
        for cell in ws['B3:B6']:
            cell[0].font = data_font
            cell[0].alignment = Alignment(horizontal='left')
        
        # Encabezados de tabla
        headers = ["No.", "Fecha", "Proveedor", "Concepto", "Folio", "RFC", "Importe"]
        row_num = 8
        
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}{row_num}"]
            cell.value = header
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = header_fill
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        # Datos de factura
        row_num = 9
        
        # Fecha en formato corto
        fecha_obj = datetime.strptime(data['Fecha_original'].split('T')[0], '%Y-%m-%d')
        fecha_corta = fecha_obj.strftime('%d/%m/%Y')
        
        # Lista de conceptos para una celda
        conceptos_texto = ", ".join([f"{desc} ({cant})" for desc, cant in data['Conceptos'].items()])
        
        # Agregar fila de datos
        ws[f"A{row_num}"] = 1
        ws[f"B{row_num}"] = fecha_corta
        ws[f"C{row_num}"] = data['Nombre_Emisor']
        ws[f"D{row_num}"] = conceptos_texto
        ws[f"E{row_num}"] = f"{data['Serie']}{data['Numero']}"
        ws[f"F{row_num}"] = data['Rfc_emisor']
        ws[f"G{row_num}"] = monto
        
        # Aplicar formato a los datos
        for col_num in range(1, 8):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}{row_num}"]
            cell.font = data_font
            cell.border = thin_border
            
            # Alineación especial para números e importes
            if col_num in [1, 7]:  # No. e Importe
                cell.alignment = Alignment(horizontal='right')
            elif col_num == 2:  # Fecha
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')
        
        # Formato de moneda para el importe
        ws[f"G{row_num}"].number_format = '"$"#,##0.00'
        
        # Total
        row_num += 2
        ws.merge_cells(f'A{row_num}:F{row_num}')
        ws[f"A{row_num}"] = "TOTAL"
        ws[f"A{row_num}"].font = header_font
        ws[f"A{row_num}"].alignment = Alignment(horizontal='right')
        
        ws[f"G{row_num}"] = monto
        ws[f"G{row_num}"].font = header_font
        ws[f"G{row_num}"].alignment = Alignment(horizontal='right')
        ws[f"G{row_num}"].number_format = '"$"#,##0.00'
        
        # Guardar el archivo
        output_path = os.path.join(output_dir, "relacion_facturas.xlsx")
        wb.save(output_path)
        
        return output_path
        
    except Exception as e:
        raise Exception(f"Error al crear relación de facturas en Excel: {str(e)}")
